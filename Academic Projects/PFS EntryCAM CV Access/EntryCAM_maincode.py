import cv2
from matplotlib import pyplot as plt
import numpy as np
import imutils
import pytesseract
import os
import pyfirmata
import openpyxl
import time
from datetime import datetime
from pyfirmata import util, Arduino
from time import process_time

# PATHS
xlDoc_path = "C:\\Users\\Ignacio V\\PycharmProjects\\pythonProject\\Registro_EntryCAM.xlsx"
valiDoc_path = "C:\\Users\\Ignacio V\\PycharmProjects\\pythonProject\\Validacion_EntryCAM.xlsx"
carImages_path = "C:\\Users\\Ignacio V\\Documents\\UMx\\PFS\\EntryCAM\\CAR_IMAGES"
plateImages_path = "C:\\Users\\Ignacio V\\Documents\\UMx\\PFS\\EntryCAM\\LICENCE_PLATE_IMAGES"
testImg_path = f"C:\\Users\\Ignacio V\\Documents\\UMx\\PFS\\EntryCAM\\Imagenes prueba"


# FUNCTIONS
def i_finder(doc_name):
    wb = openpyxl.load_workbook(doc_name)
    sheet = wb.active
    ai = 2
    while True:
        if sheet[f"A{ai}"].value is None:
            break
        else:
            ai += 1
            continue
    return ai


def take_picture():
    global i
    os.chdir(carImages_path)
    cam_port = 0 # set 1 or 2 depending on camara 0 for webcam
    cam = cv2.VideoCapture(cam_port, cv2.CAP_DSHOW)
    result, image = cam.read()
    if result:
        cv2.imwrite(f'CAR_IMAGE{i - 2}.png', image)
        return image
    else:
        print("WARNING!: NO IMAGE DETECTED")
    cv2.destroyAllWindows()


def read_image(img):
    global i
    directory = plateImages_path
    os.chdir(directory)
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    plt.imshow(cv2.cvtColor(gray, cv2.COLOR_BGR2RGB))

    # Apply Filter
    bfilter = cv2.bilateralFilter(gray, 11, 17, 17)  # Noise reduction
    edged = cv2.Canny(bfilter, 30, 200)  # Edge detection
    plt.imshow(cv2.cvtColor(edged, cv2.COLOR_BGR2RGB))
    # plt.show()
    # Find Contours

    keypoints = cv2.findContours(edged.copy(), cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
    contours = imutils.grab_contours(keypoints)
    contours = sorted(contours, key=cv2.contourArea, reverse=True)[:10]
    location = None
    for contour in contours:
        # approximate the contour
        peri = cv2.arcLength(contour, True)
        approx = cv2.approxPolyDP(contour, 0.050 * peri, True)
        # I Changed the value to .050 in order to detect plates better
        # if our approximated contour has four points, then we can assume that we have found our screen
        if len(approx) == 4:
            location = approx
            break

    # print(location)

    mask = np.zeros(gray.shape, np.uint8)
    try:
        new_image = cv2.drawContours(mask, [location], 0, 255, -1)
        new_image = cv2.bitwise_and(img, img, mask=mask)

        plt.imshow(cv2.cvtColor(new_image, cv2.COLOR_BGR2RGB))

        (x, y) = np.where(mask == 255)
        (x1, y1) = (np.min(x), np.min(y))
        (x2, y2) = (np.max(x), np.max(y))
        cropped_image = gray[x1:x2 + 1, y1:y2 + 1]

        plt.imshow(cv2.cvtColor(cropped_image, cv2.COLOR_BGR2RGB))
        # plt.show()

        pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
        result = pytesseract.image_to_string(cropped_image)

        cv2.imwrite(f'LICENCE_PLATE{i - 2}.png', cropped_image)

        if result:
            #print(result)
            return result
        else:
            print("WARNING!: NOT ABLE TO READ IMAGE")
            return "*****"

    except cv2.error:
        print("WARNING!: OPERATION UNSUCCESSFUL")
        return "!ERROR!"


def save_plate(plate):
    global i
    wb = openpyxl.load_workbook(xlDoc_path)
    sheet = wb.active
    while True:
        try:
            open(xlDoc_path, "a+")  # or "a+"
            break
        except IOError:
            print("WARNING:FILE IS OPEN, KILLING EXCEL.EXE...")
            os.system('TASKKILL /F /IM excel.exe')
            time.sleep(1)
            break
    current_time = datetime.now()
    sheet[f"A{i}"] = current_time.strftime("%D")
    sheet[f"B{i}"] = current_time.strftime("%H:%M:%S")
    sheet[f"C{i}"] = plate
    sheet[f"D{i}"] = f'=HYPERLINK("{plateImages_path}\\LICENCE_PLATE{i - 2}.png", "PLACA_{i - 2}")'
    wb.save(xlDoc_path)
    print("DATA REGISTERED")
    return 0


def valiplate(plate):
    rowtodelete = 0
    wb = openpyxl.load_workbook(valiDoc_path)
    sheet = wb["Registro de Residentes"]
    for row in sheet.rows:
        if str(row[0].value) in plate:
            for cell in row:
                print("\nACCESS:", sheet.cell(row=cell.row, column=2).value)
                return True
    sheet2 = wb["Acceso Temporal"]
    for row in sheet2.rows:
        if str(row[0].value) in plate:
            for cell in row:
                print("\nTOKEN ACCESS:", sheet2.cell(row=cell.row, column=2).value)
                rowtodelete = cell.row
                break
    while True:
        try:
            open(xlDoc_path, "a+")  # or "a+"
            break
        except IOError:
            print("WARNING: File open killing excel...")
            os.system('TASKKILL /F /IM excel.exe')
            time.sleep(1)
            break
    if rowtodelete != 0:
        sheet2.delete_rows(rowtodelete, 1)
        wb.save(valiDoc_path)
        return True


board = Arduino('COM5')
pyfirmata.util.Iterator(board).start()
# INPUTS
pres_sensor = board.get_pin("d:2:i")
pres_sensor.enable_reporting()
# OUTPUTS
relay = board.get_pin("d:3:o")
relay.write(False)
print("ENTRYCAM SYSTEM OPERATIONAL")
# Arduino LOOP
state2 = pres_sensor.read()
TEST = 1 # Delete this
while True:
    state1 = pres_sensor.read()
    if state1 != state2 and pres_sensor.read() is False:
        # print(pres_sensor.read())
        t1_start = process_time()
        i = i_finder(xlDoc_path)
        img = cv2.imread(testImg_path + f"\\image{TEST}.jpeg")
        print(f'image{TEST}.jpeg') # Delete this
        currentplate = read_image(img)
        #currentplate = read_image(take_picture())
        print(currentplate)
        save_plate(currentplate)
        if valiplate(currentplate) is True:
            relay.write(True)
            while pres_sensor.read() == state1:
                time.sleep(3)
            relay.write(False)
        t1_stop = process_time()
        print("Processing Time:", (t1_stop - t1_start) * 1000, "ms")
        TEST = TEST + 1

        if TEST == 6:
            TEST = 1

        time.sleep(2)
    state2 = state1

